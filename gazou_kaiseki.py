
import os
import tkinter

from PIL import Image, ImageTk
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
import glob
import time

######################################
#初期設定
######################################

#シリアル通信設定
COM="COM8"
bitRate=9600
#ser = serial.Serial(COM, bitRate)

#キャプチャ設定
ration = 2

######################################
#ステージ制御関数
######################################




file_path = r"E:\takasaki\master\0207"
    


print("次に動かすステージ移動範囲の入力を行って頂きます。")
x_range = 125
z_range = 100
print("ありがとうございました。")
print("")
print("---指定された条件に従って実験を開始します。---")


#####################################

experiment_number = 1


for number in range(experiment_number):
    #new_file_name = str(number) + "回目の位相板なし結果データ"
    new_file_path = file_path

    x_list = list()
    max_list = list()
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    file_name = "Without_data No." + str(number)
    sheet.title = "test_sheet_1"
    all = new_file_path + "\\" + file_name
    wb.save(all +".xlsx")
    glob.glob("*.xlsx")

    book = openpyxl.load_workbook(all + ".xlsx")
    sheet = book[sheet.title]

    for i in range(x_range):
        print("{0}番目データの{1}フォルダ".format(str(number), str(i)))
    
        folder = "x[{:0=3}".format(int(i+1))+"]"
        X_folder = new_file_path + "\\" + folder
    
        x_list.append(i)
        lumi_list = list()
    
        sheet.cell(row=i+1,column=1).value = i+1
    
        for j in range(z_range):
            print("{0}番目データの{1}フォルダのデータ{2}".format(str(number), str(i), str(j)))    
            name = "[x,z]=[{:0=3}".format(int(i+1))+","+"{:0=3}".format(int(j+1))+"]"
            r = X_folder + "\\" + name
            image = Image.open(r + ".png")
            size = image.size
            image = image.convert("L")
            array_image = np.asarray(image)
            Glay_P = 0
            for x in range(0,size[0]):
                for y in range(0,size[1]):
                    Glay_P = Glay_P + image.getpixel((x,y))
        
            lumi_list.append(Glay_P)
        sheet.cell(row=i+1,column=2).value = lumi_list.index(max(lumi_list))
        max_list.append(lumi_list.index(max(lumi_list)))
        book.save(all + ".xlsx")
    plt.plot(x_list, max_list)
    plt.show()
print("---解析終了---")
print("お疲れさまでした。")
