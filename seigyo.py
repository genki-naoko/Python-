# -*- coding: utf-8 -*-
"""
Created on Mon Jan 25 12:58:54 2021

@author: 高崎元希
"""


# -*- coding: utf-8 -*-
"""
Created on Wed Jan 20 11:02:18 2021

@author: 高崎元希
"""
import serial
import os
import tkinter
import pyautogui
from PIL import Image, ImageTk
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
import glob
import time

######################################
#初期設定
######################################

#シリアル通信設定
COM="COM7"
bitRate=9600
ser = serial.Serial(COM, bitRate)

#キャプチャ設定
ration = 2

######################################
#ステージ制御関数
######################################

#完了チェック
def BorR():
    ser.write(bytes(b"!:\r\n"))
    line = ser.readline()

    while(line == b'B\r\n'):
        ser.write(bytes(b"!:\r\n"))
        line = ser.readline()

    
#####################################
#相対移動

def RelativeMove(axis,amount):
    if(axis == "x"):
        axis = str(2)
    elif(axis == "z"):
        axis = str(1)
    cmd = ("M:"+axis+"+P"+str(amount)+"\r\n").encode()
    print(cmd)
    ser.write(cmd)
    ser.readline()
    ser.write(b"G:\r\n")
    ser.readline()
    
def RelativeMoveMin(axis,amount):
    if(axis == "x"):
        axis = str(2)
    elif(axis == "z"):
        axis = str(1)
    cmd = ("M:"+axis+"-P"+str(amount)+"\r\n").encode()
    print(cmd)
    ser.write(cmd)
    ser.readline()
    ser.write(b"G:\r\n")
    ser.readline()
    
#####################################
#現在位置取得
    
def GetPosition():
    ser.write(b"Q:\r\n")
    out = ser.readline().decode()
    print(out)

#####################################
#カメラ制御関数
#####################################

def start(event):
    global start_x, start_y
    
    canvas1.delete("rect1")
    
    canvas1.create_rectangle(event.x,event.y,event.x+1,event.y+1,outline="red",tag="rect1")
    
    start_x = event.x
    start_y = event.y
    
def drawing(event):
    
    if event.x < 0:
        end_x = 0
    else:
        end_x = min(img_resized.width,event.x)
    if event.y < 0:
        end_y = 0
    else:
        end_y = min(img_resized.width,event.y)
        
    canvas1.coords("rect1",start_x,start_y,end_x,end_y)
    
def release_action(event):
    
    start_x,start_y,end_x,end_y = [round(n*ration) for n in canvas1.coords("rect1")]
    
    pyautogui.alert("start_x:" + str(start_x) + "\n" + "start_y:" + str(start_y) + "\n" +
                    "end_x:" + str(end_x) + "\n" + "end_y:" + str(end_y))
    
#####################################
#実験開始


file_path = r"C:\Users\_\Desktop\Myschool\Lab\2022 11\1202\WIthout3kai"
    
print("撮影する範囲を赤枠で囲ってください。次で得られた座標を入力して頂きます。")
  
if __name__ =="__main__":
    img = pyautogui.screenshot()
    
    img_resized = img.resize(size = (int(img.width / ration),
                             int(img.height / ration)),
                             resample = Image.BILINEAR)
    root = tkinter.Tk()
    root.attributes("-topmost",True)
    
    img_tk = ImageTk.PhotoImage(img_resized,master = root)
    
    canvas1 = tkinter.Canvas(root,bg = "black",
                             width = img_resized.width,
                             height = img_resized.height)
    canvas1.create_image(0,0,image = img_tk, anchor = tkinter.NW)
    
    canvas1.pack()
    canvas1.bind("<ButtonPress-1>",start)
    canvas1.bind("<Button1-Motion>",drawing)
    canvas1.bind("<ButtonRelease-1>",release_action)
    
    root.mainloop()
    
####################################
    
#キャプチャ範囲の設定
print("  得られた座標を順に入力してください.")
x_1 = int(input("   xの初期座標 → "))
y_1 = int(input("   ｙの初期座標 → "))
x_2 = int(input("   xの最終座標 → "))
y_2 = int(input("   yの最終座標 → "))

print("_______________________")

print("次に動かすステージ移動範囲の入力を行って頂きます。")
x_range = 50
z_range = 100
print("ありがとうございました。")
print("")
print("---指定された条件に従って実験を開始します。---")


#####################################

experiment_number = 3
for number  in range(experiment_number):
    new_file_name = str(number) + "回目の位相板なし結果データ"
    new_file_path = file_path + "\\" + new_file_name
    os.makedirs(new_file_path, exist_ok=True)

    for i in range(x_range):
        RelativeMove("x",1)
        BorR()
        GetPosition()
        folder = "x[{:0=3}".format(int(i+1))+"]"
        X_folder = new_file_path + "\\" + folder
        os.makedirs(X_folder, exist_ok=True)


        for j in range(z_range):
            RelativeMove("z",1)
            BorR()
            name = "[x,z]=[{:0=3}".format(int(i+1))+","+"{:0=3}".format(int(j+1))+"]"
            s = pyautogui.screenshot(region = (x_1, y_1, x_2-x_1, y_2-y_1))
            r = X_folder + "\\" + name
            s.save(r + ".png")
            time.sleep(0.2)
        GetPosition()
        BorR()
        RelativeMoveMin("z", z_range)
        BorR()
    time.sleep(10)
    RelativeMoveMin("x", x_range)
    BorR()



ser.close()



for number in range(experiment_number):
    new_file_name = str(number) + "回目の位相板なし結果データ"
    new_file_path = file_path + "\\" + new_file_name

    x_list = list()
    max_list = list()
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    file_name = "Without_data No." + str(number)
    sheet.title = "test_sheet_1"
    all = new_file_path + "\\" + file_name
    wb.save(all +".xlsx")
    glob.glob("*.xlsx")

    book = openpyxl.load_workbook(all + ".xlsx")
    sheet = book[sheet.title]

    for i in range(x_range):
        print("{0}番目データの{1}フォルダ".format(str(number), str(i)))
    
        folder = "x[{:0=3}".format(int(i+1))+"]"
        X_folder = new_file_path + "\\" + folder
    
        x_list.append(i)
        lumi_list = list()
    
        sheet.cell(row=i+1,column=1).value = i+1
    
        for j in range(z_range):
            print("{0}番目データの{1}フォルダのデータ{2}".format(str(number), str(i), str(j)))    
            name = "[x,z]=[{:0=3}".format(int(i+1))+","+"{:0=3}".format(int(j+1))+"]"
            r = X_folder + "\\" + name
            image = Image.open(r + ".png")
            size = image.size
            image = image.convert("L")
            array_image = np.asarray(image)
            Glay_P = 0
            for x in range(0,size[0]):
                for y in range(0,size[1]):
                    Glay_P = Glay_P + image.getpixel((x,y))
        
            lumi_list.append(Glay_P)
        sheet.cell(row=i+1,column=2).value = lumi_list.index(max(lumi_list))
        max_list.append(lumi_list.index(max(lumi_list)))
        book.save(all + ".xlsx")

############################################
